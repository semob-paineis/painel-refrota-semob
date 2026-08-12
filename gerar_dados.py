#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 GERAR_DADOS.PY  —  Automação de Atualização do Painel REFROTA
 Ministério das Cidades / SEMOB
================================================================================

O QUE ESTE SCRIPT FAZ
---------------------
Lê a planilha mensal "Dados_Refrota_Contratações.xlsx", extrai automaticamente
todos os indicadores dos 3 cenários (Consolidado, Refrota Público, Refrota
Privado), valida a consistência dos números e gera o arquivo "dados.json" que
alimenta o painel HTML.

COMO USAR (resumo — o passo a passo completo está no MANUAL_OPERACIONAL.md)
--------------------------------------------------------------------------
    python gerar_dados.py

    # ou apontando uma planilha específica:
    python gerar_dados.py --planilha "caminho/para/Dados_Refrota_Contratações.xlsx"

REQUISITOS
----------
    pip install openpyxl

    Para recalcular as fórmulas dos 3 cenários automaticamente (recomendado),
    é necessário ter o LibreOffice instalado. Caso não haja, o script usa os
    valores já calculados que estiverem salvos na planilha e avisa sobre isso.

================================================================================
"""

import argparse
import json
import math
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
    FUSO_BRASILIA = ZoneInfo("America/Sao_Paulo")
except ImportError:
    # Python < 3.9 sem zoneinfo: fallback fixo de UTC-3 (sem horário de verão,
    # que o Brasil não usa desde 2019 — seguro para este caso).
    from datetime import timezone, timedelta
    FUSO_BRASILIA = timezone(timedelta(hours=-3))


def agora_brasilia():
    """Hora atual no fuso de Brasília — usado em todos os timestamps do
    script para que a data exibida no painel sempre bata com o horário
    real de Brasília, independente do fuso do servidor/máquina que roda
    este script."""
    return datetime.now(FUSO_BRASILIA)

try:
    import openpyxl
except ImportError:
    print("ERRO: a biblioteca 'openpyxl' não está instalada.")
    print("      Instale com:  pip install openpyxl")
    sys.exit(1)


# ----------------------------------------------------------------------------
# CONFIGURAÇÃO — mapeamento da planilha (aba "Dados para Painel")
# ----------------------------------------------------------------------------
# Se a estrutura da planilha mudar no futuro, ajuste APENAS estas constantes.

ABA_PAINEL = "Dados para Painel"
CELULA_CENARIO = "B3"           # célula do filtro (lista suspensa)

# Rótulos aceitos na célula B3 para cada cenário lógico do painel
CENARIOS = {
    "consolidado": {"b3": "Consolidado",      "rotulo": "Consolidado",      "destaque": None},
    "publico":     {"b3": "Refrota",          "rotulo": "Refrota Público",  "destaque": 0},
    "privado":     {"b3": "Refrota Privado",  "rotulo": "Refrota Privado",  "destaque": 1},
}

# Donut de veículos — linha do "Total" e colunas (Elétricos, Euro VI, Trilhos)
LINHA_VEIC_CONTRATADOS = 13
LINHA_VEIC_SELECIONADOS = 32
COL_ELETRICOS, COL_EURO6, COL_TRILHOS = 5, 6, 7   # E, F, G

# Veículos ENTREGUES — bloco K5:P13 da aba "Dados para Painel".
# Mesma estrutura dos contratados, porém em outras colunas: a linha 13 traz os
# totais e as colunas N/O/P trazem Elétricos / Euro VI / Veículos sobre trilhos.
LINHA_VEIC_ENTREGUES = 13
COL_ENT_ELETRICOS, COL_ENT_EURO6, COL_ENT_TRILHOS = 14, 15, 16   # N, O, P

# Tabelas de Região e Ano (contratados) — usadas nas tabelas e mini-gráficos
# Colunas: B(nome) C(propostas) D(veiculos) H(investimento)
REGIOES_LINHAS = range(8, 13)     # Norte..Sul (exclui Total)
ANOS_LINHAS = range(18, 22)       # 2023..2026 (exclui Total)
COL_NOME, COL_PROPOSTAS, COL_VEICULOS, COL_INVEST = 2, 3, 4, 8

# Mesmas tabelas, porém para projetos SELECIONADOS. Mesmas colunas
# (B/C/D/H), apenas em outro bloco de linhas da aba "Dados para Painel".
REGIOES_SEL_LINHAS = range(27, 32)   # Norte..Sul (exclui Total, linha 32)
ANOS_SEL_LINHAS = range(38, 42)      # 2023..2026 (exclui Total, linha 42)

# Totais de "Projetos" — alimentam os 2 donuts de projetos (texto central,
# segmentos, tooltips e legendas). Linhas de Total das tabelas de Selecionados
# (27-32) e Contratados (8-13); C=propostas, D=veiculos, H=investimento.
LINHA_TOTAL_PROJETOS_SELECIONADOS = 32
LINHA_TOTAL_PROJETOS_CONTRATADOS = 13

# Status dos projetos selecionados (compõe o donut "Projetos Selecionados"):
# Contratados / Em preparação / Desistências (A Cancelar) / Cancelados
# IMPORTANTE: a quantidade fica na coluna C, mas o INVESTIMENTO (R$) fica na
# coluna E — são colunas diferentes, não confundir (bug já corrigido uma vez).
LINHA_STATUS_CONTRATADOS = 49
LINHA_STATUS_EM_PREPARACAO = 50
LINHA_STATUS_A_CANCELAR = 51
LINHA_STATUS_CANCELADOS = 52
COL_STATUS_QTD = 3    # coluna C (quantidade de projetos)
COL_STATUS_VALOR = 5  # coluna E (investimento em R$)

# Projetos selecionados por frente (Público/Privado) — legendas dos donuts
LINHA_FRENTE_PUBLICO = 58
LINHA_FRENTE_PRIVADO = 60
COL_FRENTE_PROPOSTAS, COL_FRENTE_INVEST = 3, 5   # C (propostas), E (investimento)

# Meta de veículos selecionados para o ano corrente (gráfico "Metas")
CELULA_META_ALVO = "J38"          # meta fixa (ex.: 5.000 unidades)
CELULA_META_REALIZADO = "J39"     # veículos selecionados no ano
CELULA_META_PERCENTUAL = "J40"    # percentual concluído

# Investimento por UF — as duas tabelas ocupam as MESMAS linhas (72–98, os 27
# estados), diferindo apenas nas colunas:
#   - Contratados : coluna B (UF)  / coluna C (investimento)
#   - Selecionados: coluna G (UF)  / coluna H (investimento)
UF_CONTRATADOS_LINHAS = range(72, 99)
UF_CONTRATADOS_COL_UF, UF_CONTRATADOS_COL_VAL = 2, 3
UF_SELECIONADOS_LINHAS = range(72, 99)
UF_SELECIONADOS_COL_UF, UF_SELECIONADOS_COL_VAL = 7, 8


# ----------------------------------------------------------------------------
# UTILITÁRIOS
# ----------------------------------------------------------------------------
def num(valor):
    """Converte para número, tratando None/vazio como 0 e arredondando floats."""
    if valor is None or valor == "":
        return 0
    if isinstance(valor, float):
        return round(valor, 2)
    return valor


def cod_uf(texto):
    """'AC - Acre' -> 'AC'."""
    if not texto:
        return None
    return str(texto).split(" - ")[0].strip()


def log(nivel, msg):
    prefixos = {"ok": "  [OK]   ", "info": "  [INFO] ", "warn": "  [AVISO]", "erro": "  [ERRO] "}
    print(prefixos.get(nivel, "        ") + " " + msg)


# ----------------------------------------------------------------------------
# RECÁLCULO DAS FÓRMULAS (via LibreOffice, quando disponível)
# ----------------------------------------------------------------------------
def tem_libreoffice():
    return shutil.which("libreoffice") is not None or shutil.which("soffice") is not None


# Caminho do utilitário oficial de recálculo (skill xlsx). Se você levar este
# script para outro ambiente, ajuste RECALC_PY para o caminho correto ou deixe
# em branco para usar o método interno de fallback (--convert-to).
RECALC_PY = os.environ.get(
    "REFROTA_RECALC_PY",
    "/mnt/skills/public/xlsx/scripts/recalc.py",
)


def recalcular_cenario(planilha_origem, valor_b3, pasta_tmp):
    """
    Cria uma cópia da planilha com B3 = valor_b3, recalcula TODAS as fórmulas
    com o LibreOffice (macro calculateAll) e devolve o caminho do arquivo
    recalculado, pronto para leitura com data_only=True.
    """
    destino = os.path.join(pasta_tmp, f"cenario_{valor_b3.replace(' ', '_')}.xlsx")
    shutil.copy(planilha_origem, destino)

    wb = openpyxl.load_workbook(destino, data_only=False)
    wb[ABA_PAINEL][CELULA_CENARIO] = valor_b3
    wb.save(destino)

    if RECALC_PY and os.path.exists(RECALC_PY):
        # Método preferido: utilitário oficial que executa a macro de recálculo.
        subprocess.run(
            [sys.executable, RECALC_PY, destino, "150"],
            check=True, capture_output=True, timeout=300,
        )
    else:
        # Fallback: conversão headless (recalcula ao abrir na maioria dos casos).
        binario = shutil.which("libreoffice") or shutil.which("soffice")
        subprocess.run(
            [binario, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", pasta_tmp, destino],
            check=True, capture_output=True, timeout=300,
        )
    return destino


# ----------------------------------------------------------------------------
# EXTRAÇÃO DE UM CENÁRIO
# ----------------------------------------------------------------------------
def extrair_cenario(caminho_xlsx, chave_cenario):
    cfg = CENARIOS[chave_cenario]
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb[ABA_PAINEL]

    def cel(r, c):
        return num(ws.cell(r, c).value)

    dados = {
        "rotulo": cfg["rotulo"],
        "destaque": cfg["destaque"],
        "veiculos": {
            "eletricos": cel(LINHA_VEIC_CONTRATADOS, COL_ELETRICOS),
            "euro6": cel(LINHA_VEIC_CONTRATADOS, COL_EURO6),
            "trilhos": cel(LINHA_VEIC_CONTRATADOS, COL_TRILHOS),
        },
        "veiculosSelecionados": {
            "eletricos": cel(LINHA_VEIC_SELECIONADOS, COL_ELETRICOS),
            "euro6": cel(LINHA_VEIC_SELECIONADOS, COL_EURO6),
            "trilhos": cel(LINHA_VEIC_SELECIONADOS, COL_TRILHOS),
        },
        "veiculosEntregues": {
            "eletricos": cel(LINHA_VEIC_ENTREGUES, COL_ENT_ELETRICOS),
            "euro6": cel(LINHA_VEIC_ENTREGUES, COL_ENT_EURO6),
            "trilhos": cel(LINHA_VEIC_ENTREGUES, COL_ENT_TRILHOS),
        },
        "regioes": [],
        "anos": [],
        "regioesSelecionados": [],
        "anosSelecionados": [],
        "ufsContratados": {},
        "ufsSelecionados": {},
    }

    for r in REGIOES_LINHAS:
        dados["regioes"].append({
            "nome": ws.cell(r, COL_NOME).value,
            "propostas": cel(r, COL_PROPOSTAS),
            "veiculos": cel(r, COL_VEICULOS),
            "investimento": cel(r, COL_INVEST),
        })

    for r in ANOS_LINHAS:
        dados["anos"].append({
            "ano": ws.cell(r, COL_NOME).value,
            "propostas": cel(r, COL_PROPOSTAS),
            "veiculos": cel(r, COL_VEICULOS),
            "investimento": cel(r, COL_INVEST),
        })

    # --- Mesmas tabelas, para projetos SELECIONADOS ---
    for r in REGIOES_SEL_LINHAS:
        dados["regioesSelecionados"].append({
            "nome": ws.cell(r, COL_NOME).value,
            "propostas": cel(r, COL_PROPOSTAS),
            "veiculos": cel(r, COL_VEICULOS),
            "investimento": cel(r, COL_INVEST),
        })

    for r in ANOS_SEL_LINHAS:
        dados["anosSelecionados"].append({
            "ano": ws.cell(r, COL_NOME).value,
            "propostas": cel(r, COL_PROPOSTAS),
            "veiculos": cel(r, COL_VEICULOS),
            "investimento": cel(r, COL_INVEST),
        })

    for r in UF_CONTRATADOS_LINHAS:
        code = cod_uf(ws.cell(r, UF_CONTRATADOS_COL_UF).value)
        if code and len(code) == 2:
            dados["ufsContratados"][code] = cel(r, UF_CONTRATADOS_COL_VAL)

    for r in UF_SELECIONADOS_LINHAS:
        code = cod_uf(ws.cell(r, UF_SELECIONADOS_COL_UF).value)
        if code and len(code) == 2:
            dados["ufsSelecionados"][code] = cel(r, UF_SELECIONADOS_COL_VAL)

    # ---- Totais dos donuts de "Projetos" (antes fixos no HTML) ----
    dados["projetos"] = {
        "selecionados": {
            "propostas": cel(LINHA_TOTAL_PROJETOS_SELECIONADOS, COL_PROPOSTAS),
            "veiculos": cel(LINHA_TOTAL_PROJETOS_SELECIONADOS, COL_VEICULOS),
            "investimento": cel(LINHA_TOTAL_PROJETOS_SELECIONADOS, COL_INVEST),
        },
        "contratados": {
            "propostas": cel(LINHA_TOTAL_PROJETOS_CONTRATADOS, COL_PROPOSTAS),
            "veiculos": cel(LINHA_TOTAL_PROJETOS_CONTRATADOS, COL_VEICULOS),
            "investimento": cel(LINHA_TOTAL_PROJETOS_CONTRATADOS, COL_INVEST),
        },
        "status": {
            "contratados": {"qtd": cel(LINHA_STATUS_CONTRATADOS, COL_STATUS_QTD), "valor": cel(LINHA_STATUS_CONTRATADOS, COL_STATUS_VALOR)},
            "emPreparacao": {"qtd": cel(LINHA_STATUS_EM_PREPARACAO, COL_STATUS_QTD), "valor": cel(LINHA_STATUS_EM_PREPARACAO, COL_STATUS_VALOR)},
            "aCancelar": {"qtd": cel(LINHA_STATUS_A_CANCELAR, COL_STATUS_QTD), "valor": cel(LINHA_STATUS_A_CANCELAR, COL_STATUS_VALOR)},
            "cancelados": {"qtd": cel(LINHA_STATUS_CANCELADOS, COL_STATUS_QTD), "valor": cel(LINHA_STATUS_CANCELADOS, COL_STATUS_VALOR)},
        },
        "frentePublico": {
            "propostas": cel(LINHA_FRENTE_PUBLICO, COL_FRENTE_PROPOSTAS),
            "investimento": cel(LINHA_FRENTE_PUBLICO, COL_FRENTE_INVEST),
        },
        "frentePrivado": {
            "propostas": cel(LINHA_FRENTE_PRIVADO, COL_FRENTE_PROPOSTAS),
            "investimento": cel(LINHA_FRENTE_PRIVADO, COL_FRENTE_INVEST),
        },
    }

    # ---- Meta de veículos selecionados no ano (antes fixa no HTML) ----
    dados["meta2026"] = {
        "meta": num(ws[CELULA_META_ALVO].value),
        "realizado": num(ws[CELULA_META_REALIZADO].value),
        "percentual": round(ws[CELULA_META_PERCENTUAL].value or 0, 4),
    }

    return dados


# ----------------------------------------------------------------------------
# VALIDAÇÕES DE QUALIDADE
# ----------------------------------------------------------------------------
def validar(dados):
    """Retorna (lista_erros, lista_avisos)."""
    erros, avisos = [], []

    c = dados["consolidado"]
    p = dados["publico"]
    v = dados["privado"]

    # 1) Reconciliação: Público + Privado ≈ Consolidado (soma de UFs)
    def soma_uf(cen, campo):
        return sum(cen[campo].values())

    for campo in ["ufsContratados", "ufsSelecionados"]:
        cons = soma_uf(c, campo)
        soma = soma_uf(p, campo) + soma_uf(v, campo)
        if abs(cons - soma) > 1.0:
            erros.append(
                f"Reconciliação {campo}: Consolidado ({cons:,.2f}) ≠ "
                f"Público+Privado ({soma:,.2f}) — diferença de {abs(cons-soma):,.2f}"
            )

    # 2) Soma das regiões = soma dos anos (mesmo universo de veículos/propostas)
    for chave, cen in dados.items():
        reg_v = sum(x["veiculos"] for x in cen["regioes"])
        ano_v = sum(x["veiculos"] for x in cen["anos"])
        if reg_v != ano_v:
            avisos.append(
                f"[{chave}] Soma de veículos por região ({reg_v}) ≠ "
                f"por ano ({ano_v})"
            )
        reg_p = sum(x["propostas"] for x in cen["regioes"])
        ano_p = sum(x["propostas"] for x in cen["anos"])
        if reg_p != ano_p:
            avisos.append(
                f"[{chave}] Soma de propostas por região ({reg_p}) ≠ "
                f"por ano ({ano_p})"
            )

    # 3) Reconciliação de propostas: Público + Privado = Consolidado
    cons_p = sum(x["propostas"] for x in c["regioes"])
    soma_p = sum(x["propostas"] for x in p["regioes"]) + sum(x["propostas"] for x in v["regioes"])
    if cons_p != soma_p:
        avisos.append(
            f"Propostas: Consolidado ({cons_p}) ≠ Público+Privado ({soma_p})"
        )

    # 4) Todas as 27 UFs presentes
    for chave, cen in dados.items():
        for campo in ["ufsContratados", "ufsSelecionados"]:
            n = len(cen[campo])
            if n != 27:
                avisos.append(f"[{chave}] {campo} tem {n} UFs (esperado 27)")

    # 5) Reconciliação dos totais de "Projetos" (donuts): Público + Privado = Consolidado
    for tipo in ["selecionados", "contratados"]:
        for campo in ["propostas", "veiculos", "investimento"]:
            cons = c["projetos"][tipo][campo]
            soma = p["projetos"][tipo][campo] + v["projetos"][tipo][campo]
            diff = abs(cons - soma)
            limite = 1.0 if campo == "investimento" else 0
            if diff > limite:
                erros.append(
                    f"Reconciliação projetos.{tipo}.{campo}: Consolidado ({cons:,.2f}) "
                    f"≠ Público+Privado ({soma:,.2f})"
                )

    # 6) Status dos projetos selecionados deve somar o total de propostas selecionadas
    for chave, cen in dados.items():
        status = cen["projetos"]["status"]
        soma_status = sum(v["qtd"] for v in status.values())
        total_sel = cen["projetos"]["selecionados"]["propostas"]
        if soma_status != total_sel:
            avisos.append(
                f"[{chave}] Soma do status dos projetos selecionados ({soma_status}) "
                f"≠ total de projetos selecionados ({total_sel})"
            )
        soma_status_valor = sum(v["valor"] for v in status.values())
        total_sel_valor = cen["projetos"]["selecionados"]["investimento"]
        if abs(soma_status_valor - total_sel_valor) > 1.0:
            avisos.append(
                f"[{chave}] Soma do valor por status ({soma_status_valor:,.2f}) "
                f"≠ investimento total selecionado ({total_sel_valor:,.2f})"
            )

    # 7) Frentes (Público/Privado) devem bater com os totais por cenário próprio
    if p["projetos"]["frentePublico"]["propostas"] != p["projetos"]["selecionados"]["propostas"]:
        avisos.append(
            "Frente Público: propostas da 'frentePublico' não coincide com o "
            "total de selecionados do cenário 'publico' — confira a planilha."
        )
    if v["projetos"]["frentePrivado"]["propostas"] != v["projetos"]["selecionados"]["propostas"]:
        avisos.append(
            "Frente Privado: propostas da 'frentePrivado' não coincide com o "
            "total de selecionados do cenário 'privado' — confira a planilha."
        )

    # 8) Meta 2026: Público + Privado = Consolidado (realizado)
    meta_cons = c["meta2026"]["realizado"]
    meta_soma = p["meta2026"]["realizado"] + v["meta2026"]["realizado"]
    if meta_cons != meta_soma:
        erros.append(
            f"Reconciliação meta2026.realizado: Consolidado ({meta_cons}) "
            f"≠ Público+Privado ({meta_soma})"
        )

    # 9) Veículos entregues: Público + Privado = Consolidado
    for campo in ["eletricos", "euro6", "trilhos"]:
        cons_e = c["veiculosEntregues"][campo]
        soma_e = p["veiculosEntregues"][campo] + v["veiculosEntregues"][campo]
        if cons_e != soma_e:
            erros.append(
                f"Reconciliação veiculosEntregues.{campo}: Consolidado ({cons_e}) "
                f"≠ Público+Privado ({soma_e})"
            )

    # 11) Regiões/anos de SELECIONADOS devem bater com os totais do cenário
    for chave, cen in dados.items():
        tot_sel = cen["projetos"]["selecionados"]
        reg_v = sum(x["veiculos"] for x in cen["regioesSelecionados"])
        ano_v = sum(x["veiculos"] for x in cen["anosSelecionados"])
        if reg_v != ano_v:
            avisos.append(
                f"[{chave}] Veículos selecionados por região ({reg_v}) "
                f"≠ por ano ({ano_v})"
            )
        if reg_v != tot_sel["veiculos"]:
            avisos.append(
                f"[{chave}] Veículos selecionados por região ({reg_v}) "
                f"≠ total de selecionados ({tot_sel['veiculos']})"
            )
        reg_p = sum(x["propostas"] for x in cen["regioesSelecionados"])
        if reg_p != tot_sel["propostas"]:
            avisos.append(
                f"[{chave}] Propostas selecionadas por região ({reg_p}) "
                f"≠ total de projetos selecionados ({tot_sel['propostas']})"
            )

    # 10) Entregues não podem superar os contratados (sanidade)
    for chave, cen in dados.items():
        tot_ent = sum(cen["veiculosEntregues"].values())
        tot_con = sum(cen["veiculos"].values())
        if tot_ent > tot_con:
            avisos.append(
                f"[{chave}] Veículos entregues ({tot_ent}) superam os contratados "
                f"({tot_con}) — confira a base."
            )

    return erros, avisos


def comparar_com_anterior(dados_novos, caminho_anterior):
    """Alerta sobre variações grandes (>40%) frente ao mês anterior."""
    if not os.path.exists(caminho_anterior):
        return []
    try:
        with open(caminho_anterior, encoding="utf-8") as f:
            ant = json.load(f)
    except Exception:
        return []

    alertas = []
    def total_uf(d, cen, campo):
        return sum(d[cen][campo].values())

    for campo in ["ufsContratados", "ufsSelecionados"]:
        novo = total_uf(dados_novos, "consolidado", campo)
        velho = total_uf(ant, "consolidado", campo)
        if velho > 0:
            var = (novo - velho) / velho * 100
            if abs(var) > 40:
                alertas.append(
                    f"Variação de {var:+.1f}% em {campo} (Consolidado) "
                    f"frente ao mês anterior — confira se é esperado."
                )
    return alertas


# ----------------------------------------------------------------------------
# BLOCO COMPARATIVO (alimenta os painéis "Evolução" e "Destaques" no HTML)
# ----------------------------------------------------------------------------
def montar_comparativo(dados_novos, caminho_anterior):
    """
    Compara a extração atual com a anterior (dados.json existente) e devolve
    um bloco pronto para o painel, com:
      - metricas: linhas da tabela comparativa (Consolidado)
      - destaques: observações em linguagem natural, geradas por regras
    Devolve None se não houver snapshot anterior (ex.: primeira execução).
    """
    if not os.path.exists(caminho_anterior):
        return None
    try:
        with open(caminho_anterior, encoding="utf-8") as f:
            ant = json.load(f)
    except Exception:
        return None
    if not all(k in ant for k in ("consolidado", "publico", "privado")):
        return None

    nc, ac = dados_novos["consolidado"], ant["consolidado"]

    def veic(cen):
        v = cen["veiculos"]
        return v["eletricos"] + v["euro6"] + v["trilhos"]

    def veic_sel(cen):
        v = cen["veiculosSelecionados"]
        return v["eletricos"] + v["euro6"] + v["trilhos"]

    metricas = [
        {"rotulo": "Projetos contratados", "formato": "int",
         "anterior": ac["projetos"]["contratados"]["propostas"],
         "atual": nc["projetos"]["contratados"]["propostas"]},
        {"rotulo": "Veículos contratados", "formato": "int",
         "anterior": veic(ac), "atual": veic(nc)},
        {"rotulo": "Investimento contratado", "formato": "moeda",
         "anterior": ac["projetos"]["contratados"]["investimento"],
         "atual": nc["projetos"]["contratados"]["investimento"]},
        {"rotulo": "Projetos selecionados", "formato": "int",
         "anterior": ac["projetos"]["selecionados"]["propostas"],
         "atual": nc["projetos"]["selecionados"]["propostas"]},
        {"rotulo": "Veículos selecionados", "formato": "int",
         "anterior": veic_sel(ac), "atual": veic_sel(nc)},
        {"rotulo": "Investimento selecionado", "formato": "moeda",
         "anterior": ac["projetos"]["selecionados"]["investimento"],
         "atual": nc["projetos"]["selecionados"]["investimento"]},
        {"rotulo": "Meta 2026 (veículos)", "formato": "int",
         "anterior": ac["meta2026"]["realizado"],
         "atual": nc["meta2026"]["realizado"]},
    ]

    # ---------------- Destaques gerados por regras ----------------
    def br(v, casas=0):
        return f"{v:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    destaques = []

    # 1) Novas contratações e de qual frente vieram
    d_contr = nc["projetos"]["contratados"]["propostas"] - ac["projetos"]["contratados"]["propostas"]
    if d_contr > 0:
        d_pub = (dados_novos["publico"]["projetos"]["contratados"]["propostas"]
                 - ant["publico"]["projetos"]["contratados"]["propostas"])
        d_priv = (dados_novos["privado"]["projetos"]["contratados"]["propostas"]
                  - ant["privado"]["projetos"]["contratados"]["propostas"])
        partes = []
        if d_priv:
            partes.append(f"{br(d_priv)} na Refrota Privado")
        if d_pub:
            partes.append(f"{br(d_pub)} na Refrota Público")
        detalhe = (" — " + " e ".join(partes)) if partes else ""
        d_veic = veic(nc) - veic(ac)
        d_inv = nc["projetos"]["contratados"]["investimento"] - ac["projetos"]["contratados"]["investimento"]
        destaques.append(
            f"Foram registradas {br(d_contr)} novas contratações{detalhe}, "
            f"somando {br(d_veic)} veículos e R$ {br(d_inv/1e6,1)} milhões em investimento."
        )
    elif d_contr < 0:
        destaques.append(
            f"O total de projetos contratados recuou em {br(abs(d_contr))} — "
            f"vale conferir na base se houve cancelamento ou ajuste."
        )

    # 2) Migração entre status (Em preparação -> Contratados)
    d_prep = nc["projetos"]["status"]["emPreparacao"]["qtd"] - ac["projetos"]["status"]["emPreparacao"]["qtd"]
    if d_contr > 0 and d_prep < 0:
        destaques.append(
            f"A carteira em preparação caiu de {br(ac['projetos']['status']['emPreparacao']['qtd'])} "
            f"para {br(nc['projetos']['status']['emPreparacao']['qtd'])} projetos, confirmando o avanço "
            f"de propostas já selecionadas para a fase de contratação."
        )

    # 3) Novos projetos selecionados
    d_sel = nc["projetos"]["selecionados"]["propostas"] - ac["projetos"]["selecionados"]["propostas"]
    if d_sel > 0:
        d_inv_sel = nc["projetos"]["selecionados"]["investimento"] - ac["projetos"]["selecionados"]["investimento"]
        destaques.append(
            f"Entraram {br(d_sel)} novos projetos selecionados, elevando o investimento "
            f"selecionado em R$ {br(d_inv_sel/1e6,1)} milhões."
        )

    # 4) Mudança no ranking regional (top 3 por veículos contratados)
    top_novo = [r["nome"] for r in sorted(nc["regioes"], key=lambda r: r["veiculos"], reverse=True)[:3]]
    top_ant = [r["nome"] for r in sorted(ac["regioes"], key=lambda r: r["veiculos"], reverse=True)[:3]]
    if top_novo != top_ant:
        entrou = [r for r in top_novo if r not in top_ant]
        saiu = [r for r in top_ant if r not in top_novo]
        if entrou and saiu:
            destaques.append(
                f"Mudança no ranking regional: {entrou[0]} entrou entre as três regiões com mais "
                f"veículos contratados, no lugar do {saiu[0]}."
            )
        else:
            destaques.append(
                f"Houve troca de posições entre as três regiões com mais veículos contratados: "
                f"agora {top_novo[0]}, {top_novo[1]} e {top_novo[2]}."
            )

    # 5) Participação de elétricos na frota contratada
    tv_n, tv_a = veic(nc), veic(ac)
    if tv_n > 0 and tv_a > 0:
        pe_n = nc["veiculos"]["eletricos"] / tv_n * 100
        pe_a = ac["veiculos"]["eletricos"] / tv_a * 100
        if abs(pe_n - pe_a) >= 0.1:
            direcao = "subiu" if pe_n > pe_a else "recuou"
            destaques.append(
                f"A participação de ônibus elétricos na frota contratada {direcao} de "
                f"{br(pe_a,1)}% para {br(pe_n,1)}%."
            )

    if not destaques:
        destaques.append("Não houve variação relevante nos indicadores em relação à atualização anterior.")

    return {
        "dataAnterior": (ant.get("_meta") or {}).get("gerado_em"),
        "dataAtual": dados_novos.get("_gerado_em"),
        "metricas": metricas,
        "destaques": destaques,
    }


# ----------------------------------------------------------------------------
# FLUXO PRINCIPAL
# ----------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Gera dados.json do painel REFROTA a partir da planilha mensal."
    )
    parser.add_argument(
        "--planilha", default="Dados_Refrota_Contratações.xlsx",
        help="Caminho da planilha .xlsx (padrão: Dados_Refrota_Contratações.xlsx)"
    )
    parser.add_argument(
        "--saida", default="dados.json",
        help="Arquivo JSON de saída (padrão: dados.json)"
    )
    parser.add_argument(
        "--sem-recalculo", action="store_true",
        help="Não recalcular com LibreOffice; usar valores já salvos por cenário "
             "(requer que a planilha tenha os 3 cenários pré-calculados — não recomendado)."
    )
    parser.add_argument(
        "--forcar", action="store_true",
        help="Gera o dados.json MESMO com erros de consistência. Use apenas para "
             "inspeção/homologação — NUNCA para publicar. Os erros continuam sendo "
             "exibidos e o arquivo recebe a marca _dados_inconsistentes."
    )
    args = parser.parse_args()

    print("=" * 72)
    print(" ATUALIZAÇÃO DO PAINEL REFROTA — geração de dados.json")
    print(" " + agora_brasilia().strftime("%d/%m/%Y %H:%M"))
    print("=" * 72)

    if not os.path.exists(args.planilha):
        log("erro", f"Planilha não encontrada: {args.planilha}")
        sys.exit(1)
    log("ok", f"Planilha encontrada: {args.planilha}")

    dados = {}

    usar_recalculo = (not args.sem_recalculo) and tem_libreoffice()
    if not usar_recalculo and not args.sem_recalculo:
        log("warn", "LibreOffice não encontrado — não será possível recalcular "
                    "os cenários automaticamente.")
        log("warn", "O script tentará ler os valores já salvos na planilha.")
        log("warn", "Para o resultado mais confiável, instale o LibreOffice.")

    with tempfile.TemporaryDirectory() as tmp:
        for chave, cfg in CENARIOS.items():
            log("info", f"Processando cenário: {cfg['rotulo']} (B3='{cfg['b3']}')")
            try:
                if usar_recalculo:
                    caminho = recalcular_cenario(args.planilha, cfg["b3"], tmp)
                else:
                    caminho = args.planilha
                dados[chave] = extrair_cenario(caminho, chave)
                log("ok", f"  {cfg['rotulo']} extraído com sucesso.")
            except subprocess.TimeoutExpired:
                log("erro", "  Tempo esgotado no recálculo (LibreOffice).")
                sys.exit(1)
            except Exception as e:
                log("erro", f"  Falha ao processar {cfg['rotulo']}: {e}")
                sys.exit(1)

    # -------- Validações --------
    print("-" * 72)
    log("info", "Executando validações de qualidade...")
    erros, avisos = validar(dados)
    alertas = comparar_com_anterior(dados, args.saida)

    for a in avisos:
        log("warn", a)
    for a in alertas:
        log("warn", a)

    if erros:
        for e in erros:
            log("erro", e)
        print("-" * 72)
        if not args.forcar:
            log("erro", "Foram encontrados ERROS de consistência. "
                        "O arquivo NÃO foi gerado.")
            log("info", "Revise a planilha e rode novamente.")
            log("info", "Para gerar assim mesmo (apenas para inspeção), use --forcar.")
            sys.exit(2)
        log("warn", "MODO --forcar: o arquivo será gerado APESAR dos erros acima.")
        log("warn", "NÃO PUBLIQUE este arquivo antes de corrigir a planilha.")

    if not avisos and not alertas:
        log("ok", "Todas as validações passaram sem avisos.")

    # -------- Metadados + gravação --------
    # O comparativo precisa ser montado ANTES de sobrescrever o dados.json,
    # pois usa o arquivo anterior como referência.
    ts_atual = agora_brasilia().isoformat(timespec="seconds")
    dados["_gerado_em"] = ts_atual
    comparativo = montar_comparativo(dados, args.saida)
    dados.pop("_gerado_em", None)

    saida = {
        "_meta": {
            "gerado_em": ts_atual,
            "planilha_origem": os.path.basename(args.planilha),
            "gerado_por": "gerar_dados.py",
            "versao_estrutura": 2,
        },
        **dados,
    }
    if comparativo:
        saida["_comparativo"] = comparativo
    if erros:
        # Marca explícita de que o arquivo foi gerado com --forcar
        saida["_dados_inconsistentes"] = erros

    with open(args.saida, "w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)

    # Cópia versionada em dados/AAAA-MM.json
    os.makedirs("dados", exist_ok=True)
    versao = os.path.join("dados", agora_brasilia().strftime("dados-%Y-%m.json"))
    shutil.copy(args.saida, versao)

    print("-" * 72)
    log("ok", f"Arquivo gerado: {args.saida}")
    log("ok", f"Cópia versionada: {versao}")

    # Resumo executivo dos números
    c = dados["consolidado"]
    tot_contr = sum(c["ufsContratados"].values())
    tot_sel = sum(c["ufsSelecionados"].values())
    veic_contr = c["veiculos"]["eletricos"] + c["veiculos"]["euro6"] + c["veiculos"]["trilhos"]
    print("-" * 72)
    print("  RESUMO (Consolidado):")
    print(f"    Investimento contratado : R$ {tot_contr/1e9:,.2f} bi")
    print(f"    Investimento selecionado: R$ {tot_sel/1e9:,.2f} bi")
    print(f"    Veículos contratados    : {veic_contr:,}")

    # ---- Prévia do texto do "Resumo Executivo" exibido no painel ----
    # O painel monta este texto sozinho a partir do dados.json, mas a prévia
    # abaixo permite conferir os números antes de publicar.
    def br(v, casas=2):
        return f"{v:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    p = c["projetos"]
    v = c["veiculos"]
    sel_floor = math.floor(p["selecionados"]["investimento"] / 1e8) / 10
    top3 = sorted(c["regioes"], key=lambda r: r["veiculos"], reverse=True)[:3]
    pct_euro6 = v["euro6"] / veic_contr * 100 if veic_contr else 0
    pct_elet = v["eletricos"] / veic_contr * 100 if veic_contr else 0

    print("-" * 72)
    print("  RESUMO EXECUTIVO (confira o texto antes de publicar):")
    print(f"    ...a soma das seleções ultrapassa R$ {br(sel_floor,1)} bilhões, das quais,")
    print(f"    foram contratadas {br(p['contratados']['propostas'],0)} propostas... O investimento contratado soma")
    print(f"    R$ {br(p['contratados']['investimento']/1e9)} bilhões para a aquisição de {br(veic_contr,0)} veículos:")
    print(f"    {br(v['euro6'],0)} ônibus Euro VI, {br(v['eletricos'],0)} ônibus elétricos e {br(v['trilhos'],0)} veículos sobre trilhos.")
    print(f"    Distribuição regional: {top3[0]['nome']} ({br(top3[0]['veiculos'],0)}), "
          f"{top3[1]['nome']} ({br(top3[1]['veiculos'],0)}), {top3[2]['nome']} ({br(top3[2]['veiculos'],0)}).")
    print(f"    Euro 6: {br(pct_euro6,1)}% da frota | Elétricos: {br(pct_elet,1)}%")
    print("=" * 72)
    log("ok", "Concluído com sucesso. Faça o upload de dados.json no servidor.")
    print("=" * 72)


if __name__ == "__main__":
    main()
